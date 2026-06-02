import io
from pathlib import Path

from src.utils.embedding import get_embeddings


def load_pdf(content: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def load_docx(content: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(content))
    return "\n".join(para.text for para in doc.paragraphs if para.text.strip())


def load_xlsx(content: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True)
    rows = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
    return "\n".join(rows)


def load_markdown(content: bytes) -> str:
    return content.decode("utf-8")


LOADERS = {
    ".pdf": load_pdf,
    ".docx": load_docx,
    ".xlsx": load_xlsx,
    ".md": load_markdown,
    ".txt": load_markdown,
}


def load_document(filename: str, content: bytes) -> str:
    ext = Path(filename).suffix.lower()
    loader = LOADERS.get(ext)
    if not loader:
        raise ValueError(f"不支持的文件类型: {ext}")
    return loader(content)


def chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> list[str]:
    if len(text) <= chunk_size:
        return [text]

    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start = end - overlap
    return chunks


async def process_document(
    filename: str,
    content: bytes,
    tenant_id: str,
    category: str,
    milvus_client,
) -> int:
    text = load_document(filename, content)
    chunks = chunk_text(text)

    if not chunks:
        return 0

    embeddings = await get_embeddings(chunks)

    milvus_client.batch_insert_knowledge(
        tenant_ids=[tenant_id] * len(chunks),
        categories=[category] * len(chunks),
        contents=chunks,
        sources=[filename] * len(chunks),
        embeddings=embeddings,
    )

    return len(chunks)
